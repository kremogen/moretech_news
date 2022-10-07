import requests
from bs4 import BeautifulSoup
import openpyxl
import os

URL = 'https://www.rbc.ru/v10/search/ajax/?project=sport&material=article&dateFrom=09.02.2018&dateTo=25.02.2018&offset=0&limit=10&query=%2B'

HEADERS = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.85 YaBrowser/21.11.1.932 Yowser/2.5 Safari/537.36',
    'accept': '*/*'}

FILE = 'results_sportrbc.xlsx'

do_you_need_text = True  # нужны ли тексты статей


def get_html(url):
    r = requests.get(url, headers=HEADERS)  # получения html-кода с сайта
    return r


def get_all_articles(url):  # получение всех статей с полностью прокрученной страницы
    articles = []
    counter = 0
    offset = url.split('&offset=')[1].split('&limit=')[0]
    while True:
        print(f'Запрос {counter // 10 + 1}')
        link = url.replace(f'&offset={offset}&limit=', f'&offset={counter}&limit=')  # составление очередной ссылки
        articles_10 = get_content((get_html(link)).text)  # получение 10-ти новых статей
        if (len(articles_10) == 0):  # проверка на конец страницы
            break
        counter += 10
        articles.extend(articles_10)

    return articles


def delete_space(string):  # удаление лишних переносов строк и пробелов
    string = string.replace('\n', ' ')
    return string.strip()


def get_text(link):  # получение текста статьи
    html = get_html(link)
    soup = BeautifulSoup(html.text, 'html.parser')
    text = ''
    all_p = soup.find_all(itemprop="articleBody")  # получение всех абзацев
    for p in all_p:
        p = delete_space(p.get_text())
        text += p

    return text


def get_content(html):  # получение 10-ти статей с сайта
    all_articles = html.split('"id"')
    articles = []
    for article in all_articles:
        if (article == '{"items":[{' or article == '{"items":[]}'):
            continue
        title = article.split('"title":"')[1].split('","photo":{"')[0]  # получение заголовка
        date = article.split('"publish_date":"')[1].split('","title":"')[0]  # получение даты публикации статьи
        anons = article.split('"anons":"')[1].split('"},{')[0]  # получение анонса
        link = article.split('"fronturl":"')[1].split('","publish_date_t":"')[0]  # получение ссылки на текст статьи
        text = ''
        if (do_you_need_text):
            text = get_text(link)

        articles.append({
            'title': title,
            'date': date,
            'anons': anons,
            'text': text
        })

    return articles


def save_file(items, path):  # сохранение данных в файл
    if (os.path.exists(path)):  # проверка на существование файла
        book = openpyxl.load_workbook(path)
        sheet = book.active
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet['A1'] = 'Название статьи'
        sheet['B1'] = 'Дата'
        sheet['C1'] = 'Анонс'
        sheet['D1'] = 'Текст статьи'

    row = sheet.max_row + 1  # дописываем в конец файла

    for item in items:
        title = item['title']
        date = item['date']
        anons = item['anons']
        text = item['text']
        sheet[row][0].value = title
        sheet[row][1].value = date
        sheet[row][2].value = anons
        sheet[row][3].value = text
        row += 1

    book.save(path)
    book.close()


def make_link(date_1, date_2):  # составление ссылки по запросу
    link = URL
    dateFrom = link.split('&dateFrom=')[1].split('&dateTo=')[0]
    dateTo = link.split('&dateTo=')[1].split('&offset=')[0]
    link = link.replace(f'&dateFrom={dateFrom}&dateTo=', f'&dateFrom=' + date_1 + '&dateTo=')
    link = link.replace(f'&dateTo={dateTo}&offset=', f'&dateTo=' + date_2 + '&offset=')
    return link


def parse():
    dates = ['09.02.2022', '25.02.2022', '09.05.2018', '25.05.2018', '23.07.2021', '08.08.2021', '23.10.2021',
             '08.11.2021']  # все исследуемые даты
    prints = {  # выводы
        'Подсчет количества статей в период с 9 февраля по 25 февраля 2018 года..': make_link(dates[0], dates[1]),
        'Подсчет количества статей в период с 9 мая по 25 мая 2018 года..': make_link(dates[2], dates[3]),
        'Подсчет количества статей в период с 23 июля по 8 августа 2021 года..': make_link(dates[4], dates[5]),
        'Подсчет количества статей в период с 23 октября по 8 ноября 2021 года..': make_link(dates[6], dates[7]),
        'Подсчет количества статей 9 февраля 2018 года..': make_link(dates[0], dates[0]),
        'Подсчет количества статей 25 февраля 2018 года..': make_link(dates[1], dates[1]),
        'Подсчет количества статей 23 июля 2021 года..': make_link(dates[4], dates[4]),
        'Подсчет количества статей 8 августа 2021 года..': make_link(dates[5], dates[5])
    }

    html = get_html(URL)
    if (html.status_code == 200):  # проверка доступа к сайту
        all_articles = []
        for printer, link in prints.items():
            print(printer)
            articles = get_all_articles(link)  # получение списка статей по очередному запросу
            all_articles.extend(articles)
            save_file(articles, FILE)  # сохранение
            print(f'Найдено {len(articles)} статей\n\n')
        print('Подсчет количества статей со словосочетанием "церемония открытия" или "церемония закрытия"..')
        articles = all_articles
        ceremony = 0
        medvedeva = 0
        zagitova = 0
        averina = 0
        romashina = 0
        counter = 1
        for article in articles:  # поиск по ключевым словам
            print(f'Исследуется статья номер {counter} из {len(articles)}')
            if ('еремони' in article['text'] and ('открыти' in article['text'] or 'закрыти' in article['text'])):
                find = [article]
                save_file(find, FILE)
                ceremony += 1
            if ('Евгени' in article['text'] and 'Медведев' in article['text']):
                find = [article]
                save_file(find, FILE)
                medvedeva += 1
            if ('Алин' in article['text'] and 'Загитов' in article['text']):
                find = [article]
                save_file(find, FILE)
                zagitova += 1
            if ('Дин' in article['text'] and 'Аверин' in article['text']):
                find = [article]
                save_file(find, FILE)
                averina += 1
            if ('Светлан' in article['text'] and 'Ромашин' in article['text']):
                find = [article]
                save_file(find, FILE)
                romashina += 1
            counter += 1
        print(f'Найдено {ceremony} статей\n\n')

        print('Подсчет количества статей со словосочетанием "Евгения Медведева"..')
        print(f'Найдено {medvedeva} статей\n\n')

        print('Подсчет количества статей со словосочетанием "Алина Загитова"..')
        print(f'Найдено {zagitova} статей\n\n')

        print('Подсчет количества статей со словосочетанием "Дина Аверина"..')
        print(f'Найдено {averina} статей\n\n')

        print('Подсчет количества статей со словосочетанием "Светлана Ромашина"..')
        print(f'Найдено {romashina} статей\n\n')


parse()
